#!/usr/bin/env python3
"""
Excel交易總結導出功能
支持每日數據追加到同一個工作表
"""

import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelTradeExporter:
    def __init__(self, filename: str = "交易總結.xlsx"):
        self.filename = filename
        self.sheet_name = "每日交易總結"
        
        # Excel樣式定義
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="366092")
        self.profit_fill = PatternFill("solid", fgColor="C6EFCE")  # 淺綠色（盈利）
        self.loss_fill = PatternFill("solid", fgColor="FFC7CE")    # 淺紅色（虧損）
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_daily_summary(self, date: str, stats: Dict) -> pd.DataFrame:
        """創建單日交易總結DataFrame"""
        
        # 基本交易統計
        basic_data = {
            '日期': date,
            '交易次數': stats.get('daily_trades', 0),
            '勝率(%)': round(stats.get('daily_win_rate', 0), 2),
        }
        
        # 收益分解數據
        profit_data = {
            '程式盈虧': round(stats.get('daily_pnl', 0), 4),
            '交易盈虧': round(stats.get('realized_pnl', 0), 4),
            '資金費收入': round(stats.get('total_funding', 0), 4),
            '正資金費': round(stats.get('positive_funding', 0), 4),
            '負資金費': round(stats.get('negative_funding', 0), 4),
            '資金費次數': stats.get('funding_count', 0),
            '手續費支出': round(stats.get('total_commission', 0), 4),
            '帳戶淨利': round(stats.get('net_profit', 0), 4),
        }
        
        # 計算額外指標
        calculated_data = {
            '理論淨利': round(basic_data.get('程式盈虧', 0) + profit_data.get('資金費收入', 0) - profit_data.get('手續費支出', 0), 4),
            '實際vs理論差異': round(profit_data.get('帳戶淨利', 0) - (basic_data.get('程式盈虧', 0) + profit_data.get('資金費收入', 0) - profit_data.get('手續費支出', 0)), 4),
            '資金費率收益率': round((profit_data.get('資金費收入', 0) / max(profit_data.get('交易盈虧', 1), 1)) * 100, 2) if profit_data.get('交易盈虧', 0) != 0 else 0,
            '手續費率': round((profit_data.get('手續費支出', 0) / max(profit_data.get('交易盈虧', 1), 1)) * 100, 2) if profit_data.get('交易盈虧', 0) != 0 else 0,
        }
        
        # 合併所有數據
        all_data = {**basic_data, **profit_data, **calculated_data}
        
        # 創建DataFrame
        df = pd.DataFrame([all_data])
        return df
    
    def load_existing_data(self) -> Optional[pd.DataFrame]:
        """載入現有的Excel數據"""
        try:
            if os.path.exists(self.filename):
                df = pd.read_excel(self.filename, sheet_name=self.sheet_name)
                return df
            return None
        except Exception as e:
            print(f"載入現有Excel數據失敗: {e}")
            return None
    
    def append_daily_data(self, date: str, stats: Dict) -> bool:
        """添加或更新每日數據"""
        try:
            # 創建新的每日總結
            new_data = self.create_daily_summary(date, stats)
            
            # 載入現有數據
            existing_data = self.load_existing_data()
            
            if existing_data is not None:
                # 檢查是否已存在該日期的記錄
                if date in existing_data['日期'].values:
                    # 更新現有記錄
                    existing_data.loc[existing_data['日期'] == date] = new_data.iloc[0]
                    final_data = existing_data
                    print(f"更新 {date} 的交易總結")
                else:
                    # 添加新記錄
                    final_data = pd.concat([existing_data, new_data], ignore_index=True)
                    print(f"添加 {date} 的交易總結")
            else:
                # 創建新文件
                final_data = new_data
                print(f"創建新Excel文件，添加 {date} 的交易總結")
            
            # 按日期排序
            final_data['日期'] = pd.to_datetime(final_data['日期'])
            final_data = final_data.sort_values('日期', ascending=False)  # 最新日期在上方
            final_data['日期'] = final_data['日期'].dt.strftime('%Y-%m-%d')
            
            # 保存到Excel
            self.save_to_excel(final_data)
            return True
            
        except Exception as e:
            print(f"添加每日數據失敗: {e}")
            return False
    
    def save_to_excel(self, df: pd.DataFrame):
        """保存DataFrame到Excel並應用格式"""
        try:
            # 創建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # 寫入數據
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 設置標題行格式
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # 調整列寬並設置格式
            column_widths = {
                'A': 12,  # 日期
                'B': 10,  # 交易次數
                'C': 10,  # 勝率
                'D': 12,  # 程式盈虧
                'E': 12,  # 交易盈虧
                'F': 12,  # 資金費收入
                'G': 10,  # 正資金費
                'H': 10,  # 負資金費
                'I': 10,  # 資金費次數
                'J': 12,  # 手續費支出
                'K': 12,  # 帳戶淨利
                'L': 12,  # 理論淨利
                'M': 12,  # 差異
                'N': 12,  # 資金費率收益率
                'O': 10,  # 手續費率
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 設置數據行格式
            for row_num in range(2, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 根據盈虧設置背景色
                    if col_num == 11:  # 帳戶淨利列
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.fill = self.profit_fill
                        elif isinstance(cell.value, (int, float)) and cell.value < 0:
                            cell.fill = self.loss_fill
            
            # 添加總計行
            self.add_summary_row(ws, df)
            
            # 保存文件
            wb.save(self.filename)
            print(f"Excel文件已保存: {self.filename}")
            
        except Exception as e:
            print(f"保存Excel文件失敗: {e}")
    
    def add_summary_row(self, ws, df: pd.DataFrame):
        """添加總計行"""
        try:
            # 計算總計
            total_trades = df['交易次數'].sum()
            avg_win_rate = df['勝率(%)'].mean()
            total_program_pnl = df['程式盈虧'].sum()
            total_realized_pnl = df['交易盈虧'].sum()
            total_funding = df['資金費收入'].sum()
            total_commission = df['手續費支出'].sum()
            total_net_profit = df['帳戶淨利'].sum()
            total_funding_count = df['資金費次數'].sum()
            
            # 添加空行
            ws.append([])
            
            # 添加總計行
            summary_row = [
                '總計',
                total_trades,
                round(avg_win_rate, 2),
                round(total_program_pnl, 4),
                round(total_realized_pnl, 4),
                round(total_funding, 4),
                '', '', # 正負資金費不需要總計
                total_funding_count,
                round(total_commission, 4),
                round(total_net_profit, 4),
                '', '', '', ''  # 其他計算列不需要總計
            ]
            
            ws.append(summary_row)
            
            # 設置總計行格式
            summary_row_num = ws.max_row
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=summary_row_num, column=col_num)
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 總淨利設置顏色
                if col_num == 11 and isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.fill = self.profit_fill
                    elif cell.value < 0:
                        cell.fill = self.loss_fill
                        
        except Exception as e:
            print(f"添加總計行失敗: {e}")
    
    def export_daily_summary(self, date: str = None, stats: Dict = None) -> bool:
        """導出每日交易總結"""
        try:
            if not date:
                date = datetime.now().strftime('%Y-%m-%d')
            
            if not stats:
                # 如果沒有提供統計數據，嘗試從profit_tracker獲取
                try:
                    from profit_tracker import ProfitTracker
                    tracker = ProfitTracker()
                    stats = tracker.get_daily_stats()
                except Exception as e:
                    print(f"無法獲取統計數據: {e}")
                    return False
            
            return self.append_daily_data(date, stats)
            
        except Exception as e:
            print(f"導出每日總結失敗: {e}")
            return False
    
    def export_historical_data(self, days: int = 30) -> bool:
        """導出歷史數據（用於初始化Excel文件）"""
        try:
            from profit_tracker import ProfitTracker
            from account_analyzer import AccountAnalyzer
            
            tracker = ProfitTracker()
            analyzer = AccountAnalyzer()
            
            # 獲取每日數據
            for i in range(days):
                target_date = datetime.now() - timedelta(days=i)
                date_str = target_date.strftime('%Y-%m-%d')
                
                try:
                    # 獲取該日的詳細統計
                    start_time = int(target_date.replace(hour=0, minute=0, second=0).timestamp() * 1000)
                    end_time = int(target_date.replace(hour=23, minute=59, second=59).timestamp() * 1000)
                    
                    # 獲取帳戶數據
                    income_history = analyzer.get_account_income_history(start_time=start_time, end_time=end_time)
                    trade_history = analyzer.get_trade_history(start_time=start_time, end_time=end_time)
                    
                    if not trade_history:  # 如果該日沒有交易，跳過
                        continue
                    
                    # 分析數據
                    realized_pnl = analyzer.calculate_realized_pnl(trade_history)
                    funding_income = analyzer.get_funding_rate_income(income_history)
                    
                    # 構建統計數據
                    daily_stats = {
                        'daily_trades': len(trade_history),
                        'daily_win_rate': (len([t for t in trade_history if float(t['realizedPnl']) > 0]) / len(trade_history)) * 100,
                        'daily_pnl': realized_pnl['total_pnl'],
                        'realized_pnl': realized_pnl['total_pnl'],
                        'total_commission': realized_pnl['total_commission'],
                        'total_funding': funding_income['total_funding'],
                        'positive_funding': funding_income.get('positive_funding', 0),
                        'negative_funding': funding_income.get('negative_funding', 0),
                        'funding_count': funding_income.get('funding_count', 0),
                        'net_profit': realized_pnl['total_pnl'] + funding_income['total_funding'] - realized_pnl['total_commission']
                    }
                    
                    # 添加到Excel
                    self.append_daily_data(date_str, daily_stats)
                    print(f"已添加 {date_str} 的歷史數據")
                    
                except Exception as e:
                    print(f"處理 {date_str} 數據時出錯: {e}")
                    continue
            
            print(f"歷史數據導出完成: {self.filename}")
            return True
            
        except Exception as e:
            print(f"導出歷史數據失敗: {e}")
            return False

def main():
    """測試函數"""
    exporter = ExcelTradeExporter()
    
    # 測試數據
    test_stats = {
        'daily_trades': 15,
        'daily_win_rate': 73.3,
        'daily_pnl': 0.1234,
        'realized_pnl': 0.0567,
        'total_commission': 0.0456,
        'total_funding': 0.1123,
        'positive_funding': 0.1345,
        'negative_funding': -0.0222,
        'funding_count': 8,
        'net_profit': 0.1234
    }
    
    # 導出測試數據
    today = datetime.now().strftime('%Y-%m-%d')
    result = exporter.export_daily_summary(today, test_stats)
    
    if result:
        print("✅ Excel導出成功！")
    else:
        print("❌ Excel導出失敗")

if __name__ == '__main__':
    main() 